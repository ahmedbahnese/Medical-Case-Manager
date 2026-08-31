from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pathlib import Path

p = Path('/home/ubuntu/upload/‎⁨يوم22يوليوبيانبحالاتالتنفسالصناعيحضاناتمستشفىالاطفالالتخصصي⁩.xlsx')
wb = load_workbook(p, data_only=False)
ws = wb[wb.sheetnames[0]]
print('sheet', ws.title, 'size', ws.max_row, ws.max_column)
print('merged', list(ws.merged_cells.ranges))
print('freeze', ws.freeze_panes)
print('orientation', ws.sheet_view.rightToLeft, 'page', ws.page_setup.orientation, ws.page_setup.paperSize, 'fit', ws.page_setup.fitToWidth, ws.page_setup.fitToHeight)
print('print_area', ws.print_area)
for col in range(1, ws.max_column + 1):
    letter = get_column_letter(col)
    print('width', letter, ws.column_dimensions[letter].width)
for r in range(1, 25):
    vals = [ws.cell(r,c).value for c in range(1, ws.max_column+1)]
    if any(v is not None for v in vals):
        print('row', r, vals)
        print('styles', [ws.cell(r,c).style_id for c in range(1, ws.max_column+1)])
for name in ['A1','C3','B4','A7','A10','A11','B11','G11','A22']:
    c=ws[name]
    print('cell', name, 'value=',repr(c.value),'font=',c.font.name,c.font.sz,c.font.bold,c.font.italic,'fill=',c.fill.fgColor.rgb,c.fill.fgColor.indexed,'align=',c.alignment.horizontal,c.alignment.vertical,c.alignment.wrap_text,'border=',c.border.left.style,c.border.right.style,c.border.top.style,c.border.bottom.style,'numfmt=',c.number_format)
